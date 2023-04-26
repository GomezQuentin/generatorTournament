import csv
import io
import random
from collections import namedtuple
from typing import List
from itertools import zip_longest

from openpyxl import Workbook, load_workbook

# https://docs.python.org/3/library/itertools.html#itertools-recipes
def grouper(iterable, n, *, incomplete='fill', fillvalue=None):
    "Collect data into non-overlapping fixed-length chunks or blocks"
    # grouper('ABCDEFG', 3, fillvalue='x') --> ABC DEF Gxx
    # grouper('ABCDEFG', 3, incomplete='strict') --> ABC DEF ValueError
    # grouper('ABCDEFG', 3, incomplete='ignore') --> ABC DEF
    
    
    args = [iter(iterable)] * n
    if incomplete == 'fill':
        return zip_longest(*args, fillvalue=fillvalue)
    if incomplete == 'strict':
        return zip(*args, strict=True)
    if incomplete == 'ignore':
        return zip(*args)
    else:
        raise ValueError('Expected fill, strict, or ignore')

Cow = namedtuple('Cow', ['key', 'name', 'owner'])

def generate_tournament(cows: List[Cow]) -> Workbook:
    cow_count = len(cows)
    if (cow_count <= 16):
        slot_count = 16
    elif (cow_count <= 32):
        slot_count = 32
    elif (cow_count <= 64):
        slot_count = 64
    elif (cow_count <= 128):
        slot_count = 128
    elif (cow_count <= 256):
        slot_count = 256
    else:
        slot_count = 512

    random.shuffle(cows)

    # Load the Excel file using openpyxl
    filename = 'tournament/tournament' + str(slot_count) + '.xlsx'
    wb = load_workbook(filename)

    # Select the worksheet you want to modify
    ws = wb['Sheet1']

    rounds = grouper(cows, 2)

    for round_idx, (cow1, cow2) in enumerate(rounds, start=1):
        for cow, cell_pos in [(cow1, f'B{round_idx * 8}'), (cow2, f'B{round_idx * 8 + 4}')]:
            if cow:
                print(f'{cell_pos} = {cow=} {type(cow)=}')
                ws[cell_pos] = f'{cow.key, cow.name, cow.owner}'

    return wb

from aiohttp import web


async def request_tournament(request):
    # WARNING: don't do that if you plan to receive large files!
    data = await request.post()

    if 'csv' not in data:
        return web.json_response({'error': 'no csv file'}, status=400)
    csv_data = data['csv']
    
    # .file contains the actual file data that needs to be stored somewhere.
    csv_file = csv_data.file

    #with open(csv_file, 'r') as f:
    reader = csv.reader(io.TextIOWrapper(csv_file))
    cows = [Cow(key, name, owner) for key, name, owner in reader]

    wb = generate_tournament(cows)
    wb_file = io.BytesIO()
    buf = io.BufferedRandom(wb_file)
    wb.save(buf)

    wb_file.seek(0)
    response = web.Response(
        body=buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': 'attachment; filename="output2.xlsx"',
        }
    )

    return response

app = web.Application()
app.add_routes([web.post('/tournament', request_tournament)])

if __name__ == '__main__':
    web.run_app(app, port=8888)
