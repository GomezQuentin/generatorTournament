import numpy as np
import random
import io
import pandas as pd

from openpyxl import Workbook, load_workbook

import csv


def random_number(start, end):
    return random.randint(start, end)

def check_str(str1, str2):
    return str1 == str2

def generate_tournament(csv_input, output_file):
    # Read CSV file
    df = pd.read_csv(csv_input, header=None)

    # convert the dataframe to a numpy array
    arr = df.to_numpy()

    # get the number of rows and columns
    #num_rows, num_cols = df.shape
    num_rows = len(arr)

    # create a boolean array with the same number of rows as the original array
    bool_col = np.array([False] * arr.shape[0])

    # horizontally stack the two arrays
    arr_with_bool = np.hstack((arr, bool_col.reshape((-1, 1))))

    # create varibales
    total_cow = num_rows-1
    if(total_cow <= 16):
        total_place = 16
    elif (total_cow <=32):
        total_place = 32
    elif (total_cow <=64):
        total_place = 64
    elif (total_cow <=128):
        total_place = 128
    elif (total_cow <=256):
        total_place = 256
    else:
        total_place = 512

    copy_cow = total_place - total_cow
    iteration = total_cow - copy_cow
    # create an empty array with the same dimension as the new_arr, because we will stack them together
    final_array= np.zeros((0, 5)) 

    i=0
    while i <= iteration/2:
        while True:
            draw_cow = random_number(0, total_cow-1)
            
            if (not arr_with_bool[draw_cow][3]):
                arr_with_bool[draw_cow][3] = True
                opponent = arr_with_bool[draw_cow][0]-1
                break

        while True:
            draw_cow = random_number(0, total_cow-1)
            if (not arr_with_bool[draw_cow][3] and not check_str(arr_with_bool[draw_cow][2], arr_with_bool[int(opponent)][2])):
                arr_with_bool[draw_cow][3] = True
                new_arr = arr_with_bool[[draw_cow, int(opponent)], :]
                # add the match number to the array
                int_col = np.array([i+1] * new_arr.shape[0])
                # horizontally stack the two arrays
                new_arr = np.hstack((new_arr, int_col.reshape((-1, 1))))
                # add first game to the final array
                final_array  = np.vstack((final_array, new_arr))
                break
        i += 1

    j=0
    while j<=total_cow:
        if(not arr_with_bool[j][3]):
            arr_with_bool[j][3] = True
            new_arr = arr_with_bool[[j, j], :]
            # add the match number to the array
            int_col = np.array([i+1] * new_arr.shape[0])
            # horizontally stack the two arrays
            new_arr = np.hstack((new_arr, int_col.reshape((-1, 1))))
            # add first game to the final array
            final_array  = np.vstack((final_array, new_arr))
            i += 1
        j += 1
    # delete boolean column
    final_array = np.delete(final_array, 3, axis=1)
    
    df = pd.DataFrame({'key': final_array[:, 0], 'name': final_array[:, 1], 'owner': final_array[:, 2], 'round': final_array[:, 3]})

    # Load the Excel file using openpyxl
    filename = 'tournament/tournament' + str(total_place) + '.xlsx'
    print(filename)
    wb = load_workbook(filename)

    # Select the worksheet you want to modify
    ws = wb['Sheet1']

    i=1
    while i<=total_place:
        # Get the data you want to add to the Excel file
        if not df[df['key'] == i].empty:
            row = df.loc[df['key'] == i, ['key', 'name', 'owner', 'round']]
            cell_value = row.values[0]
            # Round value is the number of the game in the tournament
            round_value = row['round'].values[0]
            # Convert the cell to string
            cell_value = ', '.join(map(str, cell_value))
            # Add the data to a specific cell
            x=round_value*8
            B = 'B'+str(x)
            #I want to check if the ws[B] is empty or not 
            if ws[B].value == None:
                ws[B] = cell_value
                # Fill out both cell and then we override the value
                A='B'+str(x+4)
                ws[A] = cell_value
            else:
                x=x+4
                B = 'B'+str(x)
                ws[B] = cell_value
        else:
            print("no data found for ", i)
        i+=1

    # Save the changes to the Excel file
    wb.save(output_file)

from aiohttp import web

async def request_tournament(request):
    # WARNING: don't do that if you plan to receive large files!
    data = await request.post()

    # if 'csv' not in data:
    #     return web.json_response({'error': 'no csv file'}, status=400)
    csv = data['csv']

    # .filename contains the name of the file in string format.
    filename = csv.filename

    # .file contains the actual file data that needs to be stored somewhere.
    csv_file = csv.file
    #content = csv_file.read()
    print(f'{filename=}')

    # Generate the tournament using the script
    wb_file = io.BytesIO()
    buf = io.BufferedRandom(wb_file)

    generate_tournament(csv_file, buf)

    wb_file.seek(0)
    response = web.Response(
        body=buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': 'attachment; filename="output2.xlsx"',
        }
    )
    #response.headers['Content-Disposition'] = 'attachment; filename="output2.xlsx"'
    #response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response
    #response = web.FileResponse(path='output2.xlsx')

app = web.Application()
app.add_routes([web.post('/tournament', request_tournament)])

if __name__ == '__main__':
    web.run_app(app, port=8880)
