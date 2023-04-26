
import io
import random
import openpyxl
from collections import namedtuple
from typing import List

from openpyxl import Workbook, load_workbook




Cow = namedtuple('Cow', ['key', 'name', 'owner', 'check'])

def generate_tournament(cows: List[Cow]) -> Workbook:    
    cow_count = len(cows)

    if(cow_count <= 16):
        total_place = 16
    elif (cow_count <=32):
        total_place = 32
    elif (cow_count <=64):
        total_place = 64
    elif (cow_count <=128):
        total_place = 128
    elif (cow_count <=256):
        total_place = 256
    else:
        total_place = 512

    copy_cow = total_place - cow_count
    iteration = cow_count - copy_cow
    final_cows = []
    counter=0
    i=0
    while i < iteration/2:
        while True:
            draw_cow = random.choice(cows)
            if not draw_cow.check:
                opponent = Cow(draw_cow.key, draw_cow.name, draw_cow.owner, True)
                cows[cows.index(draw_cow)] = opponent
                break
        while True:
            draw_cow = random.choice(cows)
            if not draw_cow.check and (draw_cow.owner != opponent.owner):
                new_cow = Cow(draw_cow.key, draw_cow.name, draw_cow.owner, True)
                cows[cows.index(draw_cow)] = new_cow
                counter+=1
                
                final_cows.append(opponent)
                final_cows.append(new_cow)
                break
        i+=1

    #create a for loop that take all the cows that have not been checked yet
    j=0
    while j<cow_count:
        if(not cows[j].check):
            final_cows.append(cows[j])
            final_cows.append(cows[j])
        j+=1


    filename = 'tournament/tournament' + str(total_place) + '.xlsx'

    wb = load_workbook(filename)
    ws = wb['Sheet1']

    x = 0
    i = 1

    x=8
    for cow in final_cows:
        cell_value = [int(cow.key), cow.name, cow.owner]
        print(cell_value)
        cell_value = ', '.join(map(str, cell_value))
        print(cell_value)
        B = 'B' + str(x)
        x+=4
        ws[B] = cell_value

    return wb


from aiohttp import web

async def request_tournament(request):
    # WARNING: don't do that if you plan to receive large files!
    data = await request.post()

    xlsx_data = data['xlsx']

    xlsx_file = xlsx_data.file

    wb = openpyxl.load_workbook(xlsx_file)

    sheet = wb.active
    cows = [Cow(*[cell.value for cell in row], False) for row in sheet.iter_rows(min_row=2)]


    wb = generate_tournament(cows)
    wb_file = io.BytesIO()
    buf = io.BufferedRandom(wb_file)
    wb.save(buf)

    wb_file.seek(0)
    response = web.Response(
        body=buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': 'attachment; filename="output2.xlsx"',
        }
    )

    return response

app = web.Application()
app.add_routes([web.post('/tournament', request_tournament)])

if __name__ == '__main__':
    web.run_app(app, port=8888)